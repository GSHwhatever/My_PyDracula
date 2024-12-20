
# -*- coding:gbk -*-
"""
http://10.64.2.100:30010/user/menu/querydatagrid?businessID=23123232，
businessID: 23123232
接口字典
"""
from openpyxl import load_workbook
from . Base_Class import Base_Class, AuthError
from . Reset_width import Reset
from datetime import datetime
import os, time, asyncio


class RL(Base_Class):

    def __init__(self, ini_path, template_excel):
        super().__init__(ini_path, template_excel)
        p = os.path.join(self.out_path, f'个人信息查询结果{time.strftime("%Y-%m-%d")}.xlsx')
        path = p if os.path.exists(p) else os.path.abspath(os.path.join(self.template_excel, '个人信息查询结果.xlsx'))
        self.wb = load_workbook(path)
        self.wb._manual_calculation = True      # 禁止自动计算
        self.ws = self.wb['总表']
        self.ws_first = self.wb['实名制缩略表']
        self.result_dic = {}
        self.Reset = Reset()
        self.task_num = 0

    async def req_jb_message(self, id_num):
        payload = {
            'searchcont': id_num,
            'type': '0'
        }
        url=self.host + '/server/cxjg/list'
        async with self.session.post(url=url, json=payload) as result:
            res = await result.json()
            status = result.status
        if status == 200:
            d = res.get('data')
            if d: 
                d = d[0]
                c = d.get('cbMsgs')
                dic = {
                    "name": d.get('aac003'),
                    "id": id_num,
                    "phone": d.get('aae005'),
                    "sex": d.get('aac004Text'),
                    "where": d.get('aac010'),
                    "info_num": len(c),
                    "info": []
                }
                for i in c:
                    b = {
                        "bxzl": i.get('aae140Text'),
                        "grbh": i.get('aac001'),
                        "jfzt": i.get('aac031Text'),
                        "cbdw": i.get('aab004'),
                        "dwbh": i.get('aab001'),
                        "sftx": d.get('aac084Text'),
                        "jgbh": i.get('aab034'),
                        "cbsf": i.get('aac066'),
                        "xzlx": i.get('aae140'),
                        "cbzt": i.get('aac031'),
                        "cbsf": i.get('aac066Text')
                    }
                    dic['info'].append(b)
            else:
                dic = {
                    "name": "",
                    "id": id_num,
                    "phone": "",
                    "sex": "",
                    "where": ""         
                }
            self.result_dic[id_num] = dic
        else:
            print(f'status:{status}')
            print(f'res:{res}')

    def save_result(self, dic):
        idcard = dic.get('id')
        value = list(dic.values())[:5]
        e_value = [None for _ in range(6)]
        bh = max([0 if cell.value == '编号' else int(cell.value) for cell in self.ws['A'] if cell.value is not None])
        info = dic.get('info')
        if info:
            for i, d in enumerate(info):
                v = list(d.values())[:7]
                if i == 0:
                    self.ws.append([bh + 1] + value + v)
                else:
                    self.ws.append(e_value + v)
        else:
            self.ws.append([bh + 1] + value)
        
        values = list(dic.values())[:4]
        bh_ = max([0 if cell.value == '编号' else int(cell.value) for cell in self.ws_first['A'] if cell.value is not None])
        age = datetime.now().year - int(idcard[6:10])
        if age in range(16, 25):
            age_ = '16-24'
        elif age in range(25, 46):
            age_ = '25-45'
        elif age in range(46, 61):
            age_ = '46-60'
        elif age > 60:
            age_ = ">60"
        elif age < 16:
            age_ = "<16"
        if info:
            info_dic = {
                "nl": age,      # 年龄
                "nld": age_,    # 年龄段
                "sftx": "否",   # 是否退休
                "cbsf": "",     # 参保身份
                "xzlx": "",     # 险种类型
                "jfzt": "",     # 缴费状态
                "cbdw": "",     # 参保单位
                "jyxs": ""      # 就业形式(推荐)
            }
            if info[0].get('sftx') == '是':
                # 退休人员
                m = info[0]
                info_dic['sftx'] = '是'
                info_dic['cbsf'] = m.get('cbsf')
                info_dic['xzlx'] = m.get('bxzl')
                info_dic['jfzt'] = m.get('jfzt')
                info_dic['cbdw'] = m.get('cbdw')
                info_dic['jyxs'] = '退休人员'
            else:
                # 过滤正常参保缴费信息
                mess_cb = [i for i in info if i.get('cbzt') == '1']
                if mess_cb:
                    # {'101': '企业职工', '102': '以个人身份参保人员', '200': '居民', '301': '公务员', '302': '事业编制人员'}
                    mess_cbsf = [i.get('cbsf') for i in mess_cb]
                    s_mess_cbsf = set(mess_cbsf)
                    mess_sort = sorted(mess_cb, key=lambda x : x['xzlx'])
                    me = mess_sort[0]
                    info_dic['cbsf'] = me.get('cbsf')
                    info_dic['xzlx'] = me.get('bxzl')
                    info_dic['jfzt'] = me.get('jfzt')
                    info_dic['cbdw'] = me.get('cbdw')
                    if len(s_mess_cbsf) == 1:
                        sf = mess_cbsf[0]
                        if sf == '企业职工':
                            info_dic['jyxs'] = '单位就业(真实)'
                        elif sf == '以个人身份参保人员':
                            info_dic['jyxs'] = '单位就业'
                        elif sf == '居民':
                            info_dic['jyxs'] = '灵活就业'
                        elif sf == '公务员' or sf == '事业编制人员':
                            info_dic['jyxs'] = '公职人员'
                    elif len(s_mess_cbsf) == 2 and (s_mess_cbsf == {'事业编制人员', '企业职工'} or s_mess_cbsf == {'公务员', '企业职工'}):
                        info_dic['jyxs'] = '单位就业(真实)'
                    else:
                        info_dic['jyxs'] = '待判断(多种缴费身份)'
                else:
                    # 不存在参保缴费信息，有暂停缴费和终止缴费的信息
                    me = info[0]
                    info_dic['cbsf'] = me.get('cbsf')
                    info_dic['xzlx'] = me.get('bxzl')
                    info_dic['jfzt'] = me.get('jfzt')
                    info_dic['cbdw'] = me.get('cbdw')
                    info_dic['jyxs'] = '待判断(暂停缴费)'

            self.ws_first.append([bh_ + 1] + values + list(info_dic.values()))
        else:
            values[-1] = '女' if int(idcard[16]) % 2 == 0 else '男'
            self.ws_first.append([bh_ + 1] + values + [age, age_, '', '', '', '', '', '灵活就业'])
    
    async def run_first(self, ids):
        tasks = [self.req_jb_message(id) for id in ids]
        self.task_num += len(tasks)
        await asyncio.gather(*tasks)

    async def run(self, ids):
        monitor = asyncio.create_task(self.monitor_tasks())
        await self.init_session()
        await self.run_first(ids)
        # ic(self.result_dic)
        await self.close_session()
        monitor.cancel()
        try:
            await monitor
        except asyncio.CancelledError:
            pass
    
    async def monitor_tasks(self):
        while True:
            # 获取当前事件循环中的所有任务
            tasks = asyncio.all_tasks()
            print(f'Current number of running tasks: {len(tasks)}')
            await asyncio.sleep(1)  # 每秒检查一次

    def main(self, ids):
        if os.path.exists(self.out_path):
            asyncio.run(self.run(ids))
            _ = [self.save_result(self.result_dic.get(id)) for id in ids]
            self.Reset.reset(self.ws)
            self.Reset.reset(self.ws_first)
            self.wb.save(os.path.join(self.out_path, f'个人信息查询结果{time.strftime("%Y-%m-%d")}.xlsx'))
        else:
            print(f'检查输出路径:{self.out_path}')


if __name__ == "__main__":
    rl = RL(ini_path='F:\Projects\SQ\My_PyDracula\config.ini', template_excel=r'F:\Projects\SQ\My_PyDracula\template_excel')
    lis = ['230304197204164426',
    '230304199802204212',
    '23030419720224481X',
    '23030419860117442X',
    '230304199707064418',
    '230304200001124816',
    '230304198902124426',
    '230304197310275429',
    '230304200009214015',
    '23030419960223461X',
    '230304198501184428',
    '230304199305014418',
    '230304200012074420',
    '231025198608122231',
    '23030419720401441X',
    '230304197005194411',
    '230304198809054419',
    '230304197008214414',
    '230304196704064814',
    '230321197012242203',
    '23030419870704421X',
    '230304197603244415',
    '230304199007054411',
    '230304198206064423',
    '230304199005084617',
    '230304197904264233',
    '230304197106104411',
    '230304199706224416',
    '230304199209104210',
    '230304197205054819',
    '230302196503285036',
    '230304196810054417',
    '230304199610284414',
    '230304199506054811',
    '230304199911074429',
    '230304197603054814',
    '230304197305144459',
    '230304198507054421',
    '230304199707034411',
    '230306199801075925',
    '23030419701008425X',
    '230304199904057030',
    '230304199307144419',
    '230304196801254417',
    '230202198206231625',
    '230304198003094817',
    '230381199207279036',
    '230307197910244246',
    '230304199605014428',
    '230304197803114420',
    '230304199403024417',
    '230304199203214419',
    '230304199505084429',
    '230304199407254818',
    '230304199609204835',
    '230304198609074810',
    '230304198507044813',
    '230304199605224812',
    '230304197102254412',
    '23030619651204423X',
    '230321198911162201',
    '230304197705174219',
    '230306199912074220',
    '230304198609144620',
    '230304199204124415',
    '231025198207232229',
    '231025198501082241',
    '230304197101154815',
    '230304198304195216',
    '230304199006084221',
    '230304199502034813',
    '230304199508094024',
    '230304198210014410',
    '230304197801114435',
    '230305197102124016',
    '230304199506074425',
    '230304200012234412',
    '230304198901087053',
    '230304198010244422',
    '230304197901014829',
    '230304200011164424',
    '230304199209214436',
    '230304199302064428',
    '230304197510094411',
    '23030219930724402X',
    '230304199711234213',
    '230304197204044811',
    '230304198503063419',
    '230303200010295418',
    '23030419990216481X',
    '230304197006134250',
    '230304197102224416',
    '220124197403253838',
    '230304198910244428',
    '230304199301034411',
    '230321199609050836',
    '230304199408264612',
    '230304199902024219',
    '230304198501254828',
    '23030419990713402X',
    '230304199812134414',
    '230306200004094926',
    '23030419920206464X',
    '230304199807224423',
    '230306197101044029',
    '230304196803244810',
    '230304199302284439',
    '23030419880501441X',
    '230304197009044410',
    '230304199307244428',
    '230304199702134413',
    '230304199304104219',
    '230304199911074445',
    '230304197203194236',
    '230304198403014425',
    '230304199509174819',
    '230304199909195213',
    '230304198207124424',
    '230304199602284414',
    '23030519800727431X',
    '230304199905184410',
    '230304197608185012',
    '230304198002274445',
    '23030419870809442X',
    '230304197606084410',
    '23030419850324464X',
    '230304198405284410',
    '230304198912064615',
    '230304198106284816',
    '230304199211264416',
    '230304199712054417',
    '230304196509014432',
    '230304199604024413',
    '230304199106294410',
    '230304198210074421',
    '230304198605054425',
    '230304198212195219',
    '230304196402194234',
    '23030419831105442X',
    '230304199302285028',
    '230304196801094417',
    '230304199906254417',
    '230304198110164825',
    '210881198610024702',
    '230304196705234416',
    '23030619800520471X',
    '230304199603244422',
    '230304198502194636',
    '230304198402044614',
    '230304196909264414',
    '230304199712234821',
    '230304199502094824',
    '230304196810254419',
    '230304199106174435',
    '230304197103184014',
    '230304198608104619',
    '230303197608246421',
    '230304198304034412',
    '230304198006124815',
    '230304198111304842',
    '230304200302275431',
    '230304199204084812',
    '23030419951022422X',
    '230304197102074817',
    '230304197106154419',
    '230306200004094942',
    '230302197510074434',
    '230304197704194437',
    '230304200006024419',
    '230304198711044423',
    '230304196908314416',
    '230304199608294410',
    '23030419770805444X',
    '23030419721204482X',
    '230304198602014428',
    '230302198802206026',
    '230127199204032610',
    '230304198803244422',
    '230304196902104039',
    '230306197511304225',
    '230302199707296814',
    '230305198207274824',
    '230303198805185415',
    '230304198201244222',
    '230304198110054810',
    '230305197208264623',
    '230304198903124428',
    '230304196911144411',
    '230304197505064410',
    '230304197308284430',
    '230304197005294420',
    '230302197006046514',
    '230304197009094418',
    '230304198709064441',
    '23030419800324482X',
    '230304196604194814',
    '220421198805124926',
    '230304197209054445',
    '230304197107055412',
    '230302197102244414',
    '230304197907184415',
    '230304197304274411',
    '230304197008134430',
    '230306199309094040',
    '230304197308234417',
    '230304197301175071',
    '23030419710422441X',
    '23030519850417431X',
    '230304197510224458',
    '230305198104195023',
    '230302199202105023',
    '230304196906254472',
    '230304197211104413',
    '230303196911296419',
    '230304196810174419',
    '230304197702175419',
    '23030419740427542X',
    '230304197308284828',
    '230304197405024411',
    '230304198309194810',
    '230304197409104224',
    '230304197507214814',
    '23030419800211402X',
    '230307198108014018',
    '230304198302134428',
    '230304199509284612',
    '230303199008056412',
    '230304198702014628',
    '230304196901144813',
    '230304197712114820',
    '230304198704204425',
    '230304198308124810',
    '230304199205044433',
    '23030419741114441X',
    '230304197003314416',
    '230304198507144822',
    '232103198310080938',
    '23030419800125424X',
    '23030419711001462X',
    '230304198808014829',
    '230304197608064819',
    '230304198304244225',
    '230304197104055417',
    '230621197411141561',
    '230303199901076627',
    '23030419810707401X',
    '230304199005074419',
    '230304198002234224',
    '23030419760817441X',
    '230304197305185023',
    '230304199003224428',
    '231026197710152113',
    '23030419790822444X',
    '230307200008294024',
    '230304199802134445',
    '230304198909264616',
    '230304198706084826',
    '230304197403124814',
    '23030419770509442X',
    '230304197203154410',
    '230304196710134612',
    '230304198609094432']
    lis2 = ['23030419910226402X',
            '230304197504134018',
            '23030419830808402X',
            '230304196304064057',
            '23030419650918401X',
            '230304196407054214',
            '230304196710024018',
            '230304199804264032',
            '230304198001094020',
            '230304198908194222',
            '23030419750808402X',
            '230304197301067046',
            '230304197712094014',
            '230304197703064016',
            '230304198309234034',
            '230304197001164012',
            '230304196510014034',
            '230304200402155218',
            '23030419940614705X',
            '230304196603154036',
            '230304198003304431',
            '230304197101065417',
            '230304196404014014',
            '230304196804084839',
            '230304197607084017',
            '230304197407094018',
            '230304198011194017',
            '230304197510164029',
            '230304197106034425',
            '230304197311175446',
            '230304197301084022',
            '230304197301214421',
            '230307197708244226',
            '230304197011074029',
            '23030419890712402X',
            '230304197601044022',
            '231024197510074269',
            '230304197901264027',
            '230304197602204024',
            '230304198305284026',
            '230304198402284028',
            '230304197711295420',
            '230304199705064019',
            '230304197910134013',
            '230304196802255411',
            '230304197508204036',
            '230304198705074036',
            '230304197510154031',
            '230304198410125414',
            '230304197703014019',
            '230304197704204068',
            '230304197205225446',
            '230304197006064029',
            '230304197402115422',
            '230304197305124028',
            '230304196909264043',
            '230304197111024846',
            '230304197103244021',
            '23030419790826402X',
            '230304198107055441',
            '230123197312161562',
            '230304197410094027',
            '230304197810064038',
            '230302197702175010',
            '230304197807264012',
            '230304197404234011',
            '230304196702134014',
            '230304198110125420',
            '230304197212155423',
            '230304197205134042',
            '230304197804184041',
            '230304197706154228',
            '230304197404084818',
            '230304197612095415',
            '230304197805105421',
            '230304198108034028',
            '230304196405184031',
            '232321198612215889',
            '230304199510124018',
            '230304198003264425',
            '23030419781014541X',
            '230304197502204414',
            '230304196607184232',
            '230304197102254017',
            '230304197501275413',
            '230231199105102951',
            '230304197810204010',
            '23030219731125423X',
            '230303197408166611',
            '230304197004144033',
            '230304196803214013',
            '230304197709165416',
            '230304198212115418',
            '230304197710194011',
            '230304198109084043',
            '230304197611044229',
            '230304197711194021',
            '211223198211172622',
            '230119197409062240',
            '230304198103024023',
            '230304197403014228',
            '230304197108175424',
            '230304198112205424',
            '230303197903314626',
            '230304198909294428',
            '230304198109174022',
            '230304198901167045',
            '230304197705154023',
            '23030419810225402X',
            '230304198103234426',
            '230304196810315023',
            '23030419741004402X',
            '230304198707044025',
            '230304198406294821',
            '230304198502084429',
            '230304198603034420',
            '230921197401202427',
            '230304197609134022',
            '230304198311074025',
            '230304197308275462',
            '23030419821016542X',
            '230304197304105423',
            '230304197609045425',
            '230304198504074021',
            '230306198512054728',
            '230304197409185423',
            '230304198206054022',
            '230304198810134029',
            '230304197302064429',
            '23030419730907542X',
            '230304197701304047',
            '230304197702134422',
            '230304197805275420',
            '230304197404094063',
            '230304197102144029',
            '230304197902184045',
            '230304197409234045',
            '230304197107114048',
            '230304197503154068',
            '230304197708124030',
            '230304197812034625',
            '230304197711267032',
            '230304198102025411',
            '230304198112224019',
            '230304197404305019',
            '230304196712074035',
            '230304197403014041',
            '23030419750225442X',
            '230304197111234042',
            '230304196803194040',
            '230304197312174023',
            '370911196510225250',
            '230107198012010644',
            '230304197111185420',
            '230304198110124209',
            '23030419860219422X',
            '230304197610155226',
            '320324197404101865',
            '23030419811008522X',
            '230304197411124048',
            '230304197501254030',
            '230304197011114211',
            '230304198109075411',
            '230304197107304036',
            '230304196809085451',
            '230304197606115416',
            '230304196307074031',
            '230302196310134450',
            '23030419780616401X',
            '230307197003294821',
            '231025197307281226',
            '230304196603244015',
            '230304198704094617',
            '230304197901015418',
            '320324197210061981',
            '230304198606124210',
            '230304197405295422',
            '230304197008215417',
            '230304197101154620',
            '230304197502274033',
            '230304197302264025',
            '230304197207075242',
            '230304197010084081',
            '230304196707194016',
            '230304197910274024',
            '230304197302264818',
            '230304197711164623',
            '230306197308254920',
            '230822197205120329',
            '230304197905095427',
            '23030419750812401X',
            '230304197102274421',
            '23030419700620401X',
            '230304198601034822',
            '230304198107225025',
            '230902199701291716',
            '230304199605314025',
            '230321199609124321',
            '230302200004265611',
            '230304197406035446',
            '230304196911105420',
            '230304196407034416',
            '230304197804185415',
            '230304198102035425',
            '230304199801107031',
            '230304196910025429',
            '230304198707244019',
            '230304197005165442',
            '230304198406094029',
            '230304197205294011',
            '230304197607014019',
            '230304197305144029',
            '230304196810185417',
            '230304197610164018',
            '23030419780313443X',
            '230304198512214012',
            '342601196702211635',
            '230304197101315420',
            '230304197102134023',
            '230304199012184616',
            '230304198708025416',
            '230304197107314015',
            '230321197505220045',
            '230304197401154032',
            '230303199811035215',
            '230304196910195217',
            '230302197309215012',
            '230304197007294037',
            '230304197606014017',
            '230304198204045413',
            '230304198905034012',
            '230304196902214027',
            '23030419910226402X']
    lis3 = ['230304199812134414', '230304200003034021']
    rl.main(lis2)
