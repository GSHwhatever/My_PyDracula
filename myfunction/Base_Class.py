# -*- coding:gbk -*-
"""
����������Ļ���
"""
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from requests.exceptions import RequestException, ConnectionError, Timeout, TooManyRedirects
from . Dict import Dict
from pprint import pprint
import re, os, aiohttp, configparser


class Base_Class:

    def __init__(self, ini_path, template_excel):
        self.session = None
        self.bh = 0
        config = configparser.ConfigParser()
        config.read(ini_path)
        self.base_path = ini_path
        self.JB_host = 'http://' + config.get('Host', 'jb_host')
        self.host = 'http://' + config.get('Host', 'host')
        out_path = config.get('Path', 'download')
        self.out_path = out_path if os.path.exists(out_path) else os.path.join(os.environ['USERPROFILE'], 'Desktop')
        self.template_excel = template_excel
        self.dic = Dict
        self.headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/109.0.0.0 Safari/537.36',
            "Access-Token": config.get('Token', 'Acces_Token')
        }
    
    async def init_session(self):
        self.session = aiohttp.ClientSession(headers=self.headers)

    async def close_session(self):
        await self.session.close()

    def is_valid_id_number(self, id_number):
        if isinstance(id_number, str):
            # �������֤�����������ʽ
            pattern = r'^\d{17}[\dXx]$'
            # ʹ��������ʽ����ƥ��
            if re.match(pattern, id_number):
                return True
            else:
                return False
            
    def read_Acces_Token(self):
        if os.path.exists(os.path.join(os.path.dirname(__file__), 'config')):
            path = os.path.join(os.path.dirname(__file__), 'config', 'Access-Token.txt')
            if os.path.exists(path):
                with open(path, 'r') as f:
                    token = f.read()
                return token
    
    def read_file(self, path, col_tag, part=None, sheet=None):
        if not col_tag:
            col_tag = 'A'
        if not os.path.exists(path):
            pprint(f'{path},Ŀ���ļ�������')
        else:
            try:
                wb = load_workbook(path)
            except ValueError as E:
                error = f'{path}�����������أ����ܴ���δ֪��ʽ�������Ƿ����ɸѡ�Ȳ���'
                print(error)
                return error
            except InvalidFileException as E:
                error = f'���ܴ���{path.split(".")[-1]}���͵�Excel�ļ�,�Ƽ����Ϊxlsx����'
                print(error)
                return error
            else:
                ws = wb.active
                dic = []
                try:
                    if sheet:
                        if sheet.isdigit():
                            ws = wb.worksheets[int(sheet) - 1]
                        else:
                            ws = wb[sheet]
                except KeyError:
                    error = "���������ƴ����޷��ҵ�ָ���Ĺ�����"
                    print(error)
                    return error
                except IndexError:
                    error = "�ṩ�Ĺ���������������ʵ�ʹ�����������"
                    print(error)
                    return error
                finally:
                    lis = [cell.value for cell in ws[col_tag]]
                    if part:
                        start = end = 0
                        l = part.split(',') if ',' in part else part.split('.')
                        if l:
                            if l[0].isdigit():
                                start = int(l[0]) - 1
                            if len(l) >= 2 and l[1].isdigit():
                                end = int(l[1]) if not start else int(l[1]) - start
                            if start:
                                lis = lis[start:]
                            if end:
                                lis = lis[:end]
                    dic = [cell for cell in lis if cell and self.is_valid_id_number(cell)]
                    return dic
        
    def main(self, *arg, **kwargs):
        try:
            self.run_id(*arg, **kwargs)
        except ConnectionError as e:
            print('\n���Ӵ���\n', '\t���Ϸɻ�����?�Ͳ�ѯ���ˣ���С��')
        except Timeout as e:
            print('\n���ӳ�ʱ��\n', '\t������')
        except TooManyRedirects as e:
            print('�ض���������ࣺ', e)
        except RequestException as e:
            print('�����쳣��', e)
        except AuthError as e:
            mess = e.args[0] if e.args else ''
            if e.code == 401 and 'Full authentication is required' in mess:
                print('\n�����֤ʧ�ܣ�\n', '\t�𱣵�¼����?�Ͳ�ѯ���ˣ���С��')
            else:
                print(e.with_traceback(e.__traceback__))
        except PermissionError as e:
            print('û���ļ�Ȩ���޷�д�룬����ļ��Ѵ����ȹر��ļ�')
        except FileNotFoundError as e:
            print(f'�����ļ�·������:{e.filename}')
        else:
            # print('�������')
            return True


class AuthError(Exception):
    """�Զ����쳣��ʾ��"""
    
    def __init__(self, message, code=None):
        super().__init__(message)
        self.code = code
    
    def __str__(self):
        if self.code:
            return f"AuthError: {self.code} - {self.args[0]}"
        else:
            return f"AuthError: {self.args[0]}"
    

if __name__ == "__main__":
    b = Base_Class()
    s = b.read_file('C:/Users/GSH/Desktop/�½� Microsoft Excel ������.xlsx', None)
    print(s)
