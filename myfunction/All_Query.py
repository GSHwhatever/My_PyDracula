# -*- coding:gbk -*-
from openpyxl import load_workbook
from . Base_Class import Base_Class, AuthError
from tqdm import tqdm
from icecream import ic
from . Reset_width import Reset
import os, time, asyncio


class JC_Query(Base_Class):

    def __init__(self, ini_path, template_excel):
        super().__init__(ini_path, template_excel)
        self.ws_dic = {
            "ryjb_info": '��Ա������Ϣ',
            "jydj_info": '��ҵ�Ǽ���Ϣ',
            "sydj_info": 'ʧҵ�Ǽ���Ϣ',
            "jycyz_info": '��ҵ��ҵ֤��Ϣ',
            "knrd_info": '��ҵ������Ա�϶���Ϣ',
            "zgzs_info": 'ְҵ�ʸ�֤����Ϣ'
        }
        self.result_dic = {}
        p = os.path.join(self.out_path, f'������Ϣ��ѯ���{time.strftime("%Y-%m-%d")}.xlsx')
        path = p if os.path.exists(p) else os.path.abspath(os.path.join(self.template_excel, '������Ϣ��ѯ���.xlsx'))
        self.wb = load_workbook(path)
        self.wb._manual_calculation = True      # ��ֹ�Զ�����
        self.Reset = Reset()
        self.task_num = 0

    def test(self):
        url = self.JB_host + '/user/menu/querydatagrid?businessID=23123232'
        payload = {
            "businessID": "23123232"
        }
        res = self.session.post(url=url, json=payload)
        print(res.json())

    async def ryjb_info(self, id_card):
        # ��Ա������Ϣ
        url = self.JB_host + '/hljjy/lpleaf6-employment/api/aa/b/a/queryAc01Page?pageNo=1'
        payload = {
            "aac002": id_card
        }
        async with self.session.post(url=url, json=payload) as result:
            res = await result.json()
            status = result.status
        if status == 200:
            data = res.get('data')
            if data:
                l = data.get('list', [])
                d = l[0] if l else {"aac147": id_card}
                if d:
                    payload = {
                        "aac001": d.get('aac001'),
                        "aac002": d.get('aac002'),
                        "aac003": d.get('aac003'),
                        "aac147": d.get('aac147')
                    }
                    dic = {
                        "���": "1",
                        "��ᱣ�Ϻ�": d.get('aac147'),
                        "����": d.get('aac003'),
                        "�Ա�": self.dic.sex_dic.get(d.get('aac004'), d.get('aac004')),
                        "����": self.dic.mz_dic.get(d.get('aac005'), d.get('aac005')),
                        "��������": d.get('aac006'),
                        "��ҵ��ҵ֤����": d.get('aac021'),
                        "��Ա���": self.dic.rylb.get(d.get('aac027'), d.get('aac027')),
                        "�Ļ��̶�": self.dic.xl_dic.get(d.get('aac011'), d.get('aac011')),
                        "�ֻ�����": d.get('aae005'),
                        "��ҵ״̬": self.dic.jyzt.get(d.get('aac016'), d.get('aac016')),
                        "������": d.get('aae019'),
                        "����ṹ": d.get('aae020'),
                        "��������": d.get('aae036', '')[:10]
                    }
                    # ic(dic, sort_dicts=False)
                    self.result_dic[id_card] = {}
                    self.result_dic.get(id_card).update({"ryjb_info": [dic]})
                    return payload
        else:
            ic(res)
            raise AuthError(message=res.get('msg'), code=status)
    
    async def jydj_info(self, payload, filter=True):
        # ��ҵ�Ǽ���Ϣ
        url = self.JB_host + '/hljjy/lpleaf6-employment/api/aa/b/b/queryVCc03Ac01jyPage?pageNo=1'
        async with self.session.post(url=url, json=payload) as result:
            res = await result.json()
            status = result.status
        if status == 200:
            # ic(res)
            l2 = res.get('list', [])
            print(f"l2-{l2}")
            if filter:
                lt = [i for i in l2 if i.get('aae100') == "1"]
                l = lt if lt else [l2[0]]
            else:
                l = l2
            lis = []
            print(f'l:{l}')
            for i, d in enumerate(l, 1):
                dic = {
                    "���": i,
                    "��ᱣ�Ϻ�": payload.get('aac147'),
                    "����": payload.get('aac003'),
                    "�Ա�": self.dic.sex_dic.get(d.get('aac004'), d.get('aac004')),
                    "֤������": self.dic.zjlx.get(d.get('aac058'), d.get('aac058')),
                    "֤������": payload.get('aac002'),
                    "��ҵ����": d.get(''),
                    "��Ч���": self.dic.yxbs.get(d.get('aae100'), d.get('aae100')),
                    "��λ����": d.get('aab004'),
                    "��λͳһ������ô���": d.get(''),
                    "��λ����": d.get(''),
                    "��ҵ��ʽ": self.dic.jyxs.get(d.get('acc318'), d.get('acc318')),
                    "��ҵ����": d.get(''),
                    "��ҵ�Ǽ�����": d.get('aae043'),
                    "��ҵ����": d.get('acc114'),
                    "��ҵ�ص�": d.get(''),
                    "��ҵ�ص���ϸ��ַ": d.get(''),
                    "��ҵְҵ�����֣�": self.dic.gzmc.get(d.get('aca111'), d.get('aca111')),
                    "��ҵ��������": d.get('aca112'),
                    "���²�ҵ���": self.dic.cylb.get(d.get('aab054'), d.get('aab054')),
                    "�Ƿ�Ǽ�ʧҵ��Ա": d.get(''),
                    "�Ƿ��ҵ������Ա": d.get(''),
                    "ע�����ڡ����徭Ӫ��": d.get(''),
                    "��Ӫס�������徭Ӫ��": d.get(''),
                    "�������ݡ�����ҵ��": d.get(''),
                    "��ͬ��ʼ����": d.get('aae030'),
                    "��ͬ��������": d.get('aae031'),
                    "��λ����": d.get(''),
                    "��λ����": d.get(''),
                    "�Ƿ�4050��Ա": self.dic.sfry.get(d.get('acc02a'), d.get('acc02a')),
                    "н�ʱ���": d.get('acb983'),
                    "��ҵ������Ա���": self.dic.rdlx.get(d.get('acc369'), d.get('acc369')),
                    "������": d.get('aae019'),
                    "�������": d.get('aae020'),
                    "��������": d.get('aae030')
                }
                lis.append(dic)
            # ic(lis, sort_dicts=False)
            if not lis:
                lis = [{"���": 1, "��ᱣ�Ϻ�": payload.get('aac147'), "����": payload.get('aac003'), "�Ա�": '', "֤������": '', "֤������": payload.get('aac002')}]
            self.result_dic.get(payload.get('aac002')).update({"jydj_info": lis})
        else:
            raise AuthError(message=res.get('msg'), code=status)

    async def sydj_info(self, payload, filter=True):
        # ʧҵ�Ǽ���Ϣ
        url = self.JB_host + '/hljjy/lpleaf6-employment/api/aa/b/b/queryVCc02Ac01Page?pageNo=1'
        async with self.session.post(url=url, json=payload) as result:
            res = await result.json()
            status = result.status
        if status == 200:
            # ic(res)
            l2 = res.get('list', [])
            if filter:
                lt = [i for i in l2 if i.get('aae100') == "1"]
                l = lt if lt else [l2[0]]
            else:
                l = l2
            lis = []
            for i, d in enumerate(l, 1):
                dic =  {
                    "���": i,
                    "ʧҵ�Ǽ�����": d.get('aae043'),
                    "��ᱣ�Ϻ���": payload.get('aac147'),
                    "����": payload.get('aac003'),
                    "�Ա�": self.dic.sex_dic.get(d.get('aac004'), d.get('aac004')),
                    "֤������": self.dic.zjlx.get(d.get('aac058'), d.get('aac058')),
                    "֤������": d.get('aac002'),
                    "ѧ��": self.dic.xl_dic.get(d.get('aac011'), d.get('aac011')),
                    "����������": d.get('aab302'),
                    "��ס������": d.get('aac313'),
                    "ʧҵʱ��": d.get('aae043'),
                    "ʧҵԭ��": self.dic.syyy.get(d.get('ajc093'), d.get('ajc093')),
                    "��Ч���": self.dic.yxbs.get(d.get('aae100'), d.get('aae100')),
                    "������": d.get('aae019'),
                    "�������": d.get('aae020'),
                    "��������": d.get('ajc090')
                }
                lis.append(dic)
            # ic(lis, sort_dicts=False)
            if not lis:
                lis = [{"���": 1, "ʧҵ�Ǽ�����": '', "��ᱣ�Ϻ���": payload.get('aac147'), "����": payload.get('aac003'),}]
            self.result_dic.get(payload.get('aac002')).update({"sydj_info": lis})
        else:
            raise AuthError(message=res.get('msg'), code=status)

    async def jycyz_info(self, payload, filter=True):
        # ��ҵ��ҵ֤��Ϣ
        url = self.JB_host + '/hljjy/lpleaf6-employment/api/aa/b/b/queryVCc0aAc01jyPage?pageNo=1'
        async with self.session.post(url=url, json=payload) as result:
            res = await result.json()
            status = result.status
        if status == 200:
            # ic(res)
            l2 = res.get('list', [])
            lis = []
            if l2:
                if filter:
                    lt = [i for i in l2 if i.get('acc0a3') != '0' and i.get('aae100') == "1"]
                    lt = sorted(lt, key=lambda x: x.get('acc341', '').replace('-', ''), reverse=True)
                    l = lt[0] if lt else l2[:1]
                else:
                    l = l2
                for i, d in enumerate(l, 1):
                    dic =  {
                        "���": i,
                        "���֤��": payload.get('aac002'),
                        "����": payload.get('aac003'),
                        "�Ա�": self.dic.sex_dic.get(d.get('aac004'), d.get('aac004')),
                        "֤����������": self.dic.sqlx.get(d.get('acc0a1'), d.get('acc0a1')),
                        "��ҵ��ҵ֤����": d.get('aac021'),
                        "ԭ֤������": "",
                        "���ű��": self.dic.ffbj.get(d.get('acc0a3'), d.get('acc0a3')),
                        "��Ч���": self.dic.yxbs.get(d.get('aae100'), d.get('aae100')),
                        "��������": d.get('acc341'),
                        "��֤����": d.get('acc342'),
                        "������": d.get('aae019'),
                        "�������": d.get('aae020'),
                        "��������": d.get('aae036', '')[:10],
                    }
                    lis.append(dic)
            # ic(lis, sort_dicts=False)
            if not lis:
                lis = [{"���": 1, "���֤��": payload.get('aac002'), "����": payload.get('aac003'), "�Ա�": "", "֤����������": "", "��ҵ��ҵ֤����": "��"}]
            self.result_dic.get(payload.get('aac002')).update({"jycyz_info": lis})
        else:
            raise AuthError(message=res.get('msg'), code=status)

    async def knrd_info(self, payload, filter=True):
        # ��ҵ������Ա�϶���Ϣ
        url = self.JB_host + '/hljjy/lpleaf6-employment/api/aa/b/b/queryCc13Page?pageNo=1'
        async with self.session.post(url=url, json=payload) as result:
            res = await result.json()
            status = result.status
        if status == 200:
            # ic(res)
            data = res.get('data')
            if data:
                yxbj = {'0': '��', '1': '��'}
                l2 = data.get('list', [])
                lis = []
                if l2:
                    if filter:
                        lt = [i for i in l2 if i.get('aae100') == "1"]
                        lt = sorted(lt, key=lambda x: x.get('acc361', '').replace('-', ''), reverse=True)
                        l = lt[0] if lt else l2[0]
                    else:
                        l = l2
                    for i, d in enumerate(l, 1):
                        dic =  {
                            "���": i,
                            "��ᱣ�Ϻ�": payload.get('aac147'),
                            "����": payload.get('aac003'),
                            "֤������": payload.get('aac002'),
                            "�Ա�": self.dic.sex_dic.get(d.get('aac004'), d.get('aac004')),
                            "��������": d.get('aac006'),
                            "סַ": d.get('aae006'),
                            "��ϵ�绰": d.get('aae139'),
                            "��Ա���": self.dic.rylb.get(d.get('aac027'), d.get('aac027')),
                            "��Ч���": yxbj.get(d.get('aae100'), d.get('aae100')),
                            "��ҵԮ����������": self.dic.rdlx.get(d.get('acc369'), d.get('acc369')),
                            "��ҵԮ�������϶�ʱ��": d.get('acc361'),
                            "��ҵԮ��������Ч�ڿ�ʼʱ��": d.get('acc362')
                        }
                        lis.append(dic)
                # ic(lis, sort_dicts=False)
                if not lis:
                    lis = [{"���": 1, "��ᱣ�Ϻ�": payload.get('aac147'), "����": payload.get('aac003'), "֤������": payload.get('aac002'), "�Ա�": "", "��������": "", "סַ": "","��ϵ�绰": "", "��Ա���": "", "��Ч���":"��"}]
                self.result_dic.get(payload.get('aac002')).update({"knrd_info": lis})
        else:
            raise AuthError(message=res.get('msg'), code=status)

    def gxxl_info(self):
        # ��Уѧ����Ϣ
        # �ݲ�֧�ִ˹���
        pass

    async def zgzs_info(self, payload, filter=True):
        # ְҵ�ʸ�֤����Ϣ
        url = self.JB_host + '/hljjy/lpleaf6-employment/api/aa/b/b/queryProfessionalCertificate'
        async with self.session.post(url=url, json=payload) as result:
            res = await result.json()
            status = result.status
        if status == 200:
            # ic(res)
            data = res.get('data')
            if data:
                l = data.get('list', [])
                l = [] if not l else l
                lis = []
                for i, d in enumerate(l, 1):
                    level = {
                        '����': 1,
                        '�м�': 2,
                        '�߼�': 3,
                        '�ؼ�': 4
                    }
                    dic =  {
                        "���": i,
                        "����": payload.get('aac003'),
                        "֤������": payload.get('aac002'),
                        "�Ա�": d.get('xb'),
                        "ְҵ��������": d.get('zymc'),
                        "ְҵ����": d.get('zyfx'),
                        "���ܵȼ�": d.get("jndj"),
                        "����֪ʶ���Գɼ�": d.get('llcj'),
                        "���ܿ��˳ɼ�": d.get('jncj'),
                        "�����ɼ�": d.get('pdcj'),
                        "�Ļ��̶�": d.get('whcd'),
                        "֤����": d.get('zsbh'),
                        "��֤����": d.get('fzrq'),
                        "ְҵ���ܼ�������": d.get('jdsz'),
                        "��֤����": d.get('fzjg'),
                        "�������ε�λ": d.get('zrdw'),
                        "�ȼ�": d.get("jndj", '').split('/')[-1].replace('����', ''),
                        "����": level.get(d.get("jndj", '').split('/')[-1].replace('����', ''))
                    }
                    lis.append(dic)
                if filter:
                    ll = sorted(lis, key=lambda x:x.get("����"), reverse=True)
                    if ll:
                        ll_d = ll[0]
                        ll_d.pop("����")
                        lis = [ll_d]
                # ic(lis, sort_dicts=False)
                if not lis:
                    lis = [{"���": 1, "����": payload.get('aac003'), "֤������": payload.get('aac002'), "�Ա�": "", "ְҵ��������": "", "ְҵ����": "", "���ܵȼ�": "",
                            "����֪ʶ���Գɼ�": "","���ܿ��˳ɼ�": "","�����ɼ�": "","�Ļ��̶�": "","֤����": "","��֤����": "","ְҵ���ܼ�������": "","��֤����": "","�������ε�λ": "","�ȼ�": "��"}]
                self.result_dic.get(payload.get('aac002')).update({"zgzs_info": lis})
        else:
            raise AuthError(message=res.get('msg'), code=status)

    def btxx_info(self):
        # ������Ϣ
        url = self.JB_host + '/hljjy/lpleaf6-employment/api/aa/b/b/queryCc15Page?pageNo=1'
        pass

    def ygba_info(self):
        # �Ͷ��ù�������Ϣ
        url = self.JB_host + '/hljjy/lpleaf6-employment/api/aa/b/b/queryBc90Page?pageNo=1'
        pass

    def write_excel(self, sheet, datas):
        # ����ѯ���д��Excel�ļ�
        ws = self.wb[sheet]
        bh = max([0 if cell.value == '���' else int(cell.value) for cell in ws['A'] if cell.value is not None])
        for i in datas:
            i["���"] = bh + 1
            ws.append(list(i.values()))

    async def run_first(self, ids):
        tasks = [self.ryjb_info(id) for id in ids]
        self.task_num += (len(tasks) * 6)
        results = await asyncio.gather(*tasks)
        return results

    async def run_end(self, results, filter):
        tasks = []
        for payload in results:
            tasks.append(self.jydj_info(payload, filter))
            tasks.append(self.sydj_info(payload, filter))
            tasks.append(self.jycyz_info(payload, filter))
            tasks.append(self.knrd_info(payload, filter))
            tasks.append(self.zgzs_info(payload, filter))
        await asyncio.gather(*tasks)

    async def run(self, ids, filter=True):
        monitor = asyncio.create_task(self.monitor_tasks())
        await self.init_session()
        result = await self.run_first(ids)
        await self.run_end(result, filter)
        # ic(self.result_dic)
        await self.close_session()
        monitor.cancel()
        try:
            await monitor
        except asyncio.CancelledError:
            pass
    
    async def monitor_tasks(self):
        while True:
            # ��ȡ��ǰ�¼�ѭ���е���������
            tasks = asyncio.all_tasks()
            print(f'Current number of running tasks: {len(tasks)}')
            await asyncio.sleep(1)  # ÿ����һ��
    
    def save(self, id):
        for k, v in self.result_dic.get(id).items():
            self.write_excel(self.ws_dic.get(k), v)
        
    def reset_format(self):
        for sheet in self.ws_dic.values():
            ws = self.wb[sheet]
            self.Reset.reset(ws)

    def main(self, ids):
        if os.path.exists(self.out_path):
            asyncio.run(self.run(ids))
            _ = [self.save(id) for id in ids]
            self.reset_format()
            self.wb.save(os.path.join(self.out_path, f'������Ϣ��ѯ���{time.strftime("%Y-%m-%d")}.xlsx'))
        else:
            print(f'������·��:{self.out_path}')

    def search(self, id):
        asyncio.run(self.run([id], filter=False))
        return self.result_dic
    

if __name__ == '__main__':
    jc = JC_Query()
    lis = ['230304197204164426',
    '230304199802204212',
    '23030419720224481X',
    '23030419860117442X',
    '230304199707064418',
    '230304200001124816',
    '230304198902124426',
    '230304197310275429',
    '230304200009214015',
    '23030419960223461X',
    '230304198501184428',
    '230304199305014418',
    '230304200012074420',
    '231025198608122231',
    '23030419720401441X',
    '230304197005194411',
    '230304198809054419',
    '230304197008214414',
    '230304196704064814',
    '230321197012242203',
    '23030419870704421X',
    '230304197603244415',
    '230304199007054411',
    '230304198206064423',
    '230304199005084617',
    '230304197904264233',
    '230304197106104411',
    '230304199706224416',
    '230304199209104210',
    '230304197205054819',
    '230302196503285036',
    '230304196810054417',
    '230304199610284414',
    '230304199506054811',
    '230304199911074429',
    '230304197603054814',
    '230304197305144459',
    '230304198507054421',
    '230304199707034411',
    '230306199801075925',
    '23030419701008425X',
    '230304199904057030',
    '230304199307144419',
    '230304196801254417',
    '230202198206231625',
    '230304198003094817',
    '230381199207279036',
    '230307197910244246',
    '230304199605014428',
    '230304197803114420',
    '230304199403024417',
    '230304199203214419',
    '230304199505084429',
    '230304199407254818',
    '230304199609204835',
    '230304198609074810',
    '230304198507044813',
    '230304199605224812',
    '230304197102254412',
    '23030619651204423X',
    '230321198911162201',
    '230304197705174219',
    '230306199912074220',
    '230304198609144620',
    '230304199204124415',
    '231025198207232229',
    '231025198501082241',
    '230304197101154815',
    '230304198304195216',
    '230304199006084221',
    '230304199502034813',
    '230304199508094024',
    '230304198210014410',
    '230304197801114435',
    '230305197102124016',
    '230304199506074425',
    '230304200012234412',
    '230304198901087053',
    '230304198010244422',
    '230304197901014829',
    '230304200011164424',
    '230304199209214436',
    '230304199302064428',
    '230304197510094411',
    '23030219930724402X',
    '230304199711234213',
    '230304197204044811',
    '230304198503063419',
    '230303200010295418',
    '23030419990216481X',
    '230304197006134250',
    '230304197102224416',
    '220124197403253838',
    '230304198910244428',
    '230304199301034411',
    '230321199609050836',
    '230304199408264612',
    '230304199902024219',
    '230304198501254828',
    '23030419990713402X',
    '230304199812134414',
    '230306200004094926',
    '23030419920206464X',
    '230304199807224423',
    '230306197101044029',
    '230304196803244810',
    '230304199302284439',
    '23030419880501441X',
    '230304197009044410',
    '230304199307244428',
    '230304199702134413',
    '230304199304104219',
    '230304199911074445',
    '230304197203194236',
    '230304198403014425',
    '230304199509174819',
    '230304199909195213',
    '230304198207124424',
    '230304199602284414',
    '23030519800727431X',
    '230304199905184410',
    '230304197608185012',
    '230304198002274445',
    '23030419870809442X',
    '230304197606084410',
    '23030419850324464X',
    '230304198405284410',
    '230304198912064615',
    '230304198106284816',
    '230304199211264416',
    '230304199712054417',
    '230304196509014432',
    '230304199604024413',
    '230304199106294410',
    '230304198210074421',
    '230304198605054425',
    '230304198212195219',
    '230304196402194234',
    '23030419831105442X',
    '230304199302285028',
    '230304196801094417',
    '230304199906254417',
    '230304198110164825',
    '210881198610024702',
    '230304196705234416',
    '23030619800520471X',
    '230304199603244422',
    '230304198502194636',
    '230304198402044614',
    '230304196909264414',
    '230304199712234821',
    '230304199502094824',
    '230304196810254419',
    '230304199106174435',
    '230304197103184014',
    '230304198608104619',
    '230303197608246421',
    '230304198304034412',
    '230304198006124815',
    '230304198111304842',
    '230304200302275431',
    '230304199204084812',
    '23030419951022422X',
    '230304197102074817',
    '230304197106154419',
    '230306200004094942',
    '230302197510074434',
    '230304197704194437',
    '230304200006024419',
    '230304198711044423',
    '230304196908314416',
    '230304199608294410',
    '23030419770805444X',
    '23030419721204482X',
    '230304198602014428',
    '230302198802206026',
    '230127199204032610',
    '230304198803244422',
    '230304196902104039',
    '230306197511304225',
    '230302199707296814',
    '230305198207274824',
    '230303198805185415',
    '230304198201244222',
    '230304198110054810',
    '230305197208264623',
    '230304198903124428',
    '230304196911144411',
    '230304197505064410',
    '230304197308284430',
    '230304197005294420',
    '230302197006046514',
    '230304197009094418',
    '230304198709064441',
    '23030419800324482X',
    '230304196604194814',
    '220421198805124926',
    '230304197209054445',
    '230304197107055412',
    '230302197102244414',
    '230304197907184415',
    '230304197304274411',
    '230304197008134430',
    '230306199309094040',
    '230304197308234417',
    '230304197301175071',
    '23030419710422441X',
    '23030519850417431X',
    '230304197510224458',
    '230305198104195023',
    '230302199202105023',
    '230304196906254472',
    '230304197211104413',
    '230303196911296419',
    '230304196810174419',
    '230304197702175419',
    '23030419740427542X',
    '230304197308284828',
    '230304197405024411',
    '230304198309194810',
    '230304197409104224',
    '230304197507214814',
    '23030419800211402X',
    '230307198108014018',
    '230304198302134428',
    '230304199509284612',
    '230303199008056412',
    '230304198702014628',
    '230304196901144813',
    '230304197712114820',
    '230304198704204425',
    '230304198308124810',
    '230304199205044433',
    '23030419741114441X',
    '230304197003314416',
    '230304198507144822',
    '232103198310080938',
    '23030419800125424X',
    '23030419711001462X',
    '230304198808014829',
    '230304197608064819',
    '230304198304244225',
    '230304197104055417',
    '230621197411141561',
    '230303199901076627',
    '23030419810707401X',
    '230304199005074419',
    '230304198002234224',
    '23030419760817441X',
    '230304197305185023',
    '230304199003224428',
    '231026197710152113',
    '23030419790822444X',
    '230307200008294024',
    '230304199802134445',
    '230304198909264616',
    '230304198706084826',
    '230304197403124814',
    '23030419770509442X',
    '230304197203154410',
    '230304196710134612',
    '230304198609094432']
    lis2 = ['230304197204164426',
    '230304199802204212',
    '23030419720224481X',
    '23030419860117442X',
    '230304199707064418',
    '230304200001124816',
    '230304198902124426',
    '230304197310275429',
    '230304200009214015',
    '23030419960223461X',
    '230304198501184428']
    jc.main(lis)