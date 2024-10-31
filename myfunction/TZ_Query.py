# -*- coding:gbk -*-
"""
̨��ʵ����������֤

passLevel ͨ���ȼ� 02:У�鲻ͨ�� 01:У��ͨ��
riskLevel ���յȼ� 03:�߷��� 02:�з��� 01:�ͷ��� 00:δ���ַ���
code ������Ϣ��ʶ  01: �����ɹ���      02:����ʧ�ܣ�У�鲻ͨ���������ؽ��!

�߷����ڲ��˲� | ��ǰ���ܴ�����Ч�ľ�ҵ��Ϣ | ����
"""
from . Base_Class import Base_Class, AuthError
from . Reset_width import Reset
from openpyxl import load_workbook
from datetime import datetime
from icecream import ic
from copy import deepcopy
import os, time, asyncio


class JB(Base_Class):

    def __init__(self, ini_path, template_excel):
        super().__init__(ini_path, template_excel)
        self.whcd_dic = self.dic.xl_dic
        # ̨�˲�ѯ���������"�������Ļ��̶ȣ���ҵ�Ǽ���Ϣ�������Ƿ��ҵ���ѣ��ȼ�֤��"
        self.TZ_dic = {
            "����": "",     # ����
            "���֤��": "",       # ���֤��
            "�绰��": "",       # �绰��
            "�Ļ��̶�": "",     # �Ļ��̶�
            "�ǼǾ�ҵʱ��": datetime.now().replace(day=1).date(),     # �ǼǾ�ҵʱ��
            "��ҵ����": "",     # ��ҵ����
            "�Ƿ�Ǽ�ʧҵ��Ա": "",     # �Ƿ�Ǽ�ʧҵ��Ա
            "�Ƿ��ҵ����": "",       # �Ƿ��ҵ����
            "�Ƿ�¼���": "��",     # �Ƿ�¼���
            "�ȼ�֤��": "",     # �ȼ�֤��
            "����": "",     # ����
            "���۾���": "��",     # ���۾���
            "�м���": "��",   # �м���
            "�Ա�": "",     # �Ա�
            "��������": "",       # ��������
            "����": "",     # ����
            "��ҵ��ҵ֤��": "",     # ��ҵ��ҵ֤��
            "�����": "",       # �����
            "��ҵ�Ǽ���": "",     # ��ҵ�Ǽ���
            "����ע��": "",     # ������ע��Ϊ��Ч���˻��߸��幤�̻���Ա�����������ҵ�� 
            "����״̬": "",     # ����״̬Ϊ����״̬�����������ҵ��
            "���ݰ���": ""      # ����Ա�Ѿ���������ҵ��
        }
        self.result_dic = {}
        p = os.path.join(self.out_path, f'̨�˲�ѯ���{time.strftime("%Y-%m-%d")}.xlsx')
        path = p if os.path.exists(p) else os.path.abspath(os.path.join(self.template_excel, '̨�˲�ѯ���.xlsx'))
        self.wb = load_workbook(path)
        self.wb._manual_calculation = True      # ��ֹ�Զ�����
        self.ws_yl = self.wb['̨��Ԥ��']
        self.ws = self.wb['̨��']
        self.Reset = Reset()
        self.task_num = 0
    
    async def req_gr_message(self, id_num, TZ_dic):
        # ������Ϣ��ѯ
        payload = {
            'aac002': id_num
        }
        url = self.JB_host + '/hljjy/lpleaf6-employment/api/aa/b/a/queryAc01Page?pageNo=1'
        async with self.session.post(url=url, json=payload) as result:
            res = await result.json()
            status = result.status
        if status == 200:
            data = res.get('data').get('list')
            age = datetime.now().year - int(id_num[6:10])
            if age in range(16, 25):
                age_ = '16-24'
            elif age in range(25, 46):
                age_ = '25-45'
            elif age in range(46, 61):
                age_ = '46-60'
            if data:
                datas = data[0]
                TZ_dic.update({"����": datas.get("aac003")})
                TZ_dic.update({"�Ա�": self.dic.sex_dic.get(datas.get("aac004"), datas.get("aac004"))})
                TZ_dic.update({"����": self.dic.mz_dic.get(datas.get("aac005"), datas.get("aac005"))})
                TZ_dic.update({"���֤��": datas.get("aac002")})
                TZ_dic.update({"�绰��": datas.get("aae005")})
                TZ_dic.update({"��������": datas.get("aae020")})
                TZ_dic.update({"����": age})
                TZ_dic.update({"�����": age_})
                TZ_dic.update({"�Ļ��̶�": self.whcd_dic.get(datas.get("aac011"), datas.get("aac011"))})
                payload = {
                    "aac001": datas.get("aac001"),
                    "aac002": datas.get("aac002"),
                    "aac003": datas.get("aac003"),
                    "aac147": datas.get("aac147")
                }
                self.result_dic.update({id_num: TZ_dic})
                return payload
            else:
                TZ_dic.update({"���֤��": id_num})
                TZ_dic.update({"����": age})
                TZ_dic.update({"�����": age_})
                self.result_dic.update({id_num: TZ_dic})
                if hasattr(res, 'json'):
                    ic(f'data_else:{res.json()}')
                else:
                    ic(f'{id_num}_res:{res}')
        else:
            raise AuthError(message=res.get('msg'), code=status)
    
    async def req_jycyz_message(self, payload):
        # ��ҵ��ҵ֤��Ϣ��ѯ
        url = self.JB_host + '/hljjy/lpleaf6-employment/api/aa/b/b/queryVCc0aAc01jyPage?pageNo=1'
        async with self.session.post(url=url, json=payload) as res:
            res_jy = await res.json()
            status = res.status
        if status == 200:
            TZ_dic = self.result_dic.get(payload.get('aac002'))
            l = res_jy.get('list', [])
            cyid = '��'
            if l:
                lt = [i for i in l if i.get('acc0a3') != '0' and i.get('aae100') == "1"]
                lt = sorted(lt, key=lambda x: x.get('acc341', '').replace('-', ''), reverse=True)
                d = lt[0] if lt else l[0]
                if d:
                    cyid = d.get('aac021')
            TZ_dic.update({"��ҵ��ҵ֤��": cyid})
        else:
            raise AuthError(message=res_jy.get('msg'), code=status)

    async def req_zyzgzs_message(self, payload):
        # ְҵ�ʸ�֤����Ϣ��ѯ
        url = self.JB_host + '/hljjy/lpleaf6-employment/api/aa/b/b/queryProfessionalCertificate'
        async with self.session.post(url=url, json=payload) as res:
            res_jy = await res.json()
            status = res.status
        if status == 200:
            TZ_dic = self.result_dic.get(payload.get('aac002'))
            data = res_jy.get('data', {})
            tag = True if data.get('totalCount') else False       # ְҵ�ʸ�֤����Ϣ�Ƿ����
            level = {
                '����': 1,
                '�м�': 2,
                '�߼�': 3,
                '�ؼ�': 4
            }
            jn_lis = [i.get("jndj", '').split('/')[-1].replace('����', '') for i in data.get("list")] if tag else ""
            lis = sorted([{"dj": i, "jb": level.get(i)} for i in jn_lis], key=lambda x:x.get("jb"), reverse=True)
            djzs = "��" if not lis else lis[0].get('dj')
            TZ_dic.update({"�ȼ�֤��": djzs})
        else:
            raise AuthError(message=res_jy.get('msg'), code=status)

    async def req_jykn_message(self, payload):
        # ��ҵ�����϶���Ϣ��ѯ
        url = self.JB_host + '/hljjy/lpleaf6-employment/api/aa/b/b/queryCc13Page?pageNo=1'
        async with self.session.post(url=url, json=payload) as res:
            res_jy = await res.json()
            status = res.status
        if status == 200:
            TZ_dic = self.result_dic.get(payload.get('aac002'))
            data = res_jy.get('data', {})
            lis = [i.get('aae100') for i in data.get('list')] if data.get('list') else []       # ��ҵ�����϶���Ϣ�Ƿ����
            jykn = "��" if '1' in lis else "��"     # �Ƿ������Ч��ҵ�����϶���Ϣ
            TZ_dic.update({"�Ƿ��ҵ����": jykn})
        else:
            raise AuthError(message=res_jy.get('msg'), code=status)

    async def res_jy_message(self, payload):
        # ��ҵ��Ϣ��ѯ
        url = self.JB_host + '/hljjy/lpleaf6-employment/api/aa/b/b/queryVCc03Ac01jyPage?pageNo=1'
        async with self.session.post(url=url, json=payload) as res:
            res_jy = await res.json()
            status = res.status
        if status == 200:
            TZ_dic = self.result_dic.get(payload.get('aac002'))
            num = res_jy.get('totalCount') if res_jy.get('totalCount') else 0       # ��ҵ��Ϣ����
            TZ_dic.update({"��ҵ�Ǽ���": num})
        else:
            raise AuthError(message=res_jy.get('msg'), code=status)

    async def res_sy_message(self, payload):
        # ʧҵ��Ϣ��ѯ
        url = self.JB_host + '/hljjy/lpleaf6-employment/api/aa/b/b/queryVCc02Ac01Page?pageNo=1'
        async with self.session.post(url=url, json=payload) as res:
            res_jy = await res.json()
            status = res.status
        if status == 200:
            TZ_dic = self.result_dic.get(payload.get('aac002'))
            jylx = '���ξ�ҵ' if not res_jy.get('totalCount') and not TZ_dic.get('��ҵ�Ǽ���') else 'ʧҵ�پ�ҵ'  # ��ҵ����
            tag = '��' if jylx == '���ξ�ҵ' else '��'
            TZ_dic.update({"�Ƿ�Ǽ�ʧҵ��Ա": tag})
            TZ_dic.update({"��ҵ����": jylx})
        else:
            raise AuthError(message=res_jy.get('msg'), code=status)

    async def req_lhjy_message(self, id_num):
        # ����ҵ�Ǽ�
        url = self.JB_host + f'/hljjy/lpleaf6-employment/api/quicksearch/queryAc01ByWzCondition?condition={id_num}'
        async with self.session.post(url=url) as result:
            res = await result.json()
            status = result.status
        if status == 200:
            if res:
                payload = res[0]
                url2 = self.JB_host + '/hljjy/lpleaf6-employment/api/ca/a/c/load'
                async with self.session.post(url=url2, json=payload) as result2:
                    res2 = await result2.json()
                    status2 = result2.status
                if status2 == 200:
                    TZ_dic = self.result_dic.get(id_num)
                    code = res2.get('code')
                    if code == '01':
                        TZ_dic['����״̬'] = 'δ���ַ���'
                        TZ_dic['���ݰ���'] = 'δ���ַ���'
                    elif code == '02':
                        dic = {"03":"�߷���", "02":"�з���", "01":"�ͷ���", "00":"δ���ַ���"}
                        l = [(i.get('riskLevel'), i.get('passLevel')) for i in res2.get('riskInfo').get('messages')  if i.get('ruleCode') in ('L0400003', 'L04020003004')]
                        txbl, sczt = l
                        if txbl[1] == '01':
                            TZ_dic['���ݰ���'] = 'ͨ��'
                        elif txbl[1] == '02':
                            TZ_dic['���ݰ���'] = dic.get(l[0])
                        if sczt[1] == '01':
                            TZ_dic['����״̬'] = 'ͨ��'
                        elif sczt[1] == '02':
                            TZ_dic['����״̬'] = dic.get(l[0])
                    else:
                        ic(f'��ҵ�Ǽ�code:{code}')
                else:
                    raise AuthError(message=res2.get('msg'), code=status2)
            else:
                ic(f'{id_num}:�����ھ�ҵ�Ǽ���Ϣ')
        else:
            raise AuthError(message=res.get('msg'), code=status)

    async def req_sy_message(self, id_num):
        # ʧҵ�Ǽ�
        url = self.JB_host + f'/hljjy/lpleaf6-employment/api/quicksearch/queryAc01ByWzCondition?condition={id_num}'
        async with self.session.post(url=url) as result:
            res = await result.json()
            status = result.status
        if status == 200:
            if res:
                payload = res[0]
                url2 = self.JB_host + '/hljjy/lpleaf6-employment/api/ca/b/a/load'
                async with self.session.post(url=url2, json=payload) as result2:
                    res2 = await result2.json()
                    status2 = result2.status
                if status2 == 200:
                    TZ_dic = self.result_dic.get(id_num)
                    code = res2.get('code')
                    if code == '01':
                        TZ_dic['����ע��'] = 'ͨ��'
                    elif code == '02':
                        l = [(i.get('riskLevel'), i.get('passLevel')) for i in res2.get('riskInfo').get('messages')  if i.get('ruleCode')=='L0400004'][0]
                        if l[1] == '01':
                            TZ_dic['����ע��'] = 'ͨ��'
                        elif l[1] == '02':
                            dic = {"03":"�߷���", "02":"�з���", "01":"�ͷ���", "00":"δ���ַ���"}
                            TZ_dic['����ע��'] = dic.get(l[0])
                    else:
                        ic(f'ʧҵ�Ǽ�code:{code}')
                else:
                    raise AuthError(message=res2.get('msg'), code=status2)
            else:
                ic(f'{id_num}:������ʧҵ�Ǽ���Ϣ')
        else:
            raise AuthError(message=res.get('msg'), code=status)
        
    def save_result(self, dic, ws):
        bh = max([0 if cell.value == '���' else int(cell.value) for cell in ws['A'] if cell.value is not None])
        self.bh = bh if bh != 0 else self.bh
        values = list(dic.values())
        values.insert(0, self.bh + 1)
        # print(f"values:{values}")
        ws.append(values)
    
    def reset_format(self, ws, hidden=False):
        self.Reset.reset(ws)
        if hidden:
            # ���ؿ���
            ws.column_dimensions.group('G', 'H', hidden=True)
            ws.column_dimensions.group('J', 'N', hidden=True)
            ws.column_dimensions.group('Y', 'AB', hidden=True)
    
    async def run_first(self, ids):
        tasks = [self.req_gr_message(id, deepcopy(self.TZ_dic)) for id in ids]
        self.task_num += (len(tasks) * 8)
        results = await asyncio.gather(*tasks)
        return results
    
    async def run_end(self, results):
        tasks = []
        for payload in results:
            if payload:
                tasks.append(self.req_zyzgzs_message(payload))
                tasks.append(self.req_jykn_message(payload))
                tasks.append(self.res_jy_message(payload))
                tasks.append(self.res_sy_message(payload))
                tasks.append(self.req_jycyz_message(payload))
                tasks.append(self.req_lhjy_message(payload.get('aac002')))
                tasks.append(self.req_sy_message(payload.get('aac002')))
        await asyncio.gather(*tasks)

    async def run(self, ids):
        monitor = asyncio.create_task(self.monitor_tasks())
        await self.init_session()
        result = await self.run_first(ids)
        await self.run_end(result)
        # ic(self.result_dic)
        await self.close_session()
        monitor.cancel()
        try:
            await monitor
        except asyncio.CancelledError:
            pass
    
    async def monitor_tasks(self):
        while True:
            # ��ȡ��ǰ�¼�ѭ���е���������
            tasks = asyncio.all_tasks()
            print(f'Current number of running tasks: {len(tasks)}')
            await asyncio.sleep(1)  # ÿ����һ��
    
    def save(self, id):
        TZ_dic = self.result_dic.get(id)
        l = ['����', '���֤��', '�绰��', '�Ļ��̶�', '�ǼǾ�ҵʱ��', '��ҵ��λ', '��ҵ��ʽ', '��ҵ����', 'ͳһ���ô���', '��֯��������', '������ҵ', '��λ��ַ', '��ҵ����', '�Ƿ�Ǽ�ʧҵ��Ա', '�Ƿ��ҵ����', '�Ƿ�¼���', '�ȼ�֤��', '����', '���۾���', '�м���', '�Ա�', '��������', '����', '���֤ȱλ', '�绰����λ��', '���֤У��', '���ô���λ��', '��ҵ��ҵ֤��', '�����']
        l_yl = ['����', '���֤��', '�绰��', '��ҵ�Ǽ���', '����ע��', '����״̬', '���ݰ���']
        dic = {i:TZ_dic.get(i) for i in l}
        dic_yl = {i:TZ_dic.get(i) for i in l_yl}
        if dic_yl:
            self.save_result(dic_yl, self.ws_yl)
        if dic:
            self.save_result(dic, self.ws)

    def main(self, ids):
        if os.path.exists(self.out_path):
            asyncio.run(self.run(ids))
            # asyncio.run(self.save(ids))
            _ = [self.save(id) for id in ids]
            self.reset_format(self.ws, hidden=True)
            self.reset_format(self.ws_yl)
            self.wb.save(os.path.join(self.out_path, f'̨�˲�ѯ���{time.strftime("%Y-%m-%d")}.xlsx'))
        else:
            print(f'������·��:{self.out_path}')


if __name__ == '__main__':
    jb = JB(ini_path='F:\Projects\SQ\My_PyDracula\config.ini', template_excel=r'F:\Projects\SQ\My_PyDracula\template_excel')
    # jb.main(r'C:\Users\GSH\Desktop\test.xlsx')
    lis = ['230304197204164426',
    '230304199802204212',
    '23030419720224481X',
    '23030419860117442X',
    '230304199707064418',
    '230304200001124816',
    '230304198902124426',
    '230304197310275429',
    '230304200009214015',
    '23030419960223461X',
    '230304198501184428',
    '230304199305014418',
    '230304200012074420',
    '231025198608122231',
    '23030419720401441X',
    '230304197005194411',
    '230304198809054419',
    '230304197008214414',
    '230304196704064814',
    '230321197012242203',
    '23030419870704421X',
    '230304197603244415',
    '230304199007054411',
    '230304198206064423',
    '230304199005084617',
    '230304197904264233',
    '230304197106104411',
    '230304199706224416',
    '230304199209104210',
    '230304197205054819',
    '230302196503285036',
    '230304196810054417',
    '230304199610284414',
    '230304199506054811',
    '230304199911074429',
    '230304197603054814',
    '230304197305144459',
    '230304198507054421',
    '230304199707034411',
    '230306199801075925',
    '23030419701008425X',
    '230304199904057030',
    '230304199307144419',
    '230304196801254417',
    '230202198206231625',
    '230304198003094817',
    '230381199207279036',
    '230307197910244246',
    '230304199605014428',
    '230304197803114420',
    '230304199403024417',
    '230304199203214419',
    '230304199505084429',
    '230304199407254818',
    '230304199609204835',
    '230304198609074810',
    '230304198507044813',
    '230304199605224812',
    '230304197102254412',
    '23030619651204423X',
    '230321198911162201',
    '230304197705174219',
    '230306199912074220',
    '230304198609144620',
    '230304199204124415',
    '231025198207232229',
    '231025198501082241',
    '230304197101154815',
    '230304198304195216',
    '230304199006084221',
    '230304199502034813',
    '230304199508094024',
    '230304198210014410',
    '230304197801114435',
    '230305197102124016',
    '230304199506074425',
    '230304200012234412',
    '230304198901087053',
    '230304198010244422',
    '230304197901014829',
    '230304200011164424',
    '230304199209214436',
    '230304199302064428',
    '230304197510094411',
    '23030219930724402X',
    '230304199711234213',
    '230304197204044811',
    '230304198503063419',
    '230303200010295418',
    '23030419990216481X',
    '230304197006134250',
    '230304197102224416',
    '220124197403253838',
    '230304198910244428',
    '230304199301034411',
    '230321199609050836',
    '230304199408264612',
    '230304199902024219',
    '230304198501254828',
    '23030419990713402X',
    '230304199812134414',
    '230306200004094926',
    '23030419920206464X',
    '230304199807224423',
    '230306197101044029',
    '230304196803244810',
    '230304199302284439',
    '23030419880501441X',
    '230304197009044410',
    '230304199307244428',
    '230304199702134413',
    '230304199304104219',
    '230304199911074445',
    '230304197203194236',
    '230304198403014425',
    '230304199509174819',
    '230304199909195213',
    '230304198207124424',
    '230304199602284414',
    '23030519800727431X',
    '230304199905184410',
    '230304197608185012',
    '230304198002274445',
    '23030419870809442X',
    '230304197606084410',
    '23030419850324464X',
    '230304198405284410',
    '230304198912064615',
    '230304198106284816',
    '230304199211264416',
    '230304199712054417',
    '230304196509014432',
    '230304199604024413',
    '230304199106294410',
    '230304198210074421',
    '230304198605054425',
    '230304198212195219',
    '230304196402194234',
    '23030419831105442X',
    '230304199302285028',
    '230304196801094417',
    '230304199906254417',
    '230304198110164825',
    '210881198610024702',
    '230304196705234416',
    '23030619800520471X',
    '230304199603244422',
    '230304198502194636',
    '230304198402044614',
    '230304196909264414',
    '230304199712234821',
    '230304199502094824',
    '230304196810254419',
    '230304199106174435',
    '230304197103184014',
    '230304198608104619',
    '230303197608246421',
    '230304198304034412',
    '230304198006124815',
    '230304198111304842',
    '230304200302275431',
    '230304199204084812',
    '23030419951022422X',
    '230304197102074817',
    '230304197106154419',
    '230306200004094942',
    '230302197510074434',
    '230304197704194437',
    '230304200006024419',
    '230304198711044423',
    '230304196908314416',
    '230304199608294410',
    '23030419770805444X',
    '23030419721204482X',
    '230304198602014428',
    '230302198802206026',
    '230127199204032610',
    '230304198803244422',
    '230304196902104039',
    '230306197511304225',
    '230302199707296814',
    '230305198207274824',
    '230303198805185415',
    '230304198201244222',
    '230304198110054810',
    '230305197208264623',
    '230304198903124428',
    '230304196911144411',
    '230304197505064410',
    '230304197308284430',
    '230304197005294420',
    '230302197006046514',
    '230304197009094418',
    '230304198709064441',
    '23030419800324482X',
    '230304196604194814',
    '220421198805124926',
    '230304197209054445',
    '230304197107055412',
    '230302197102244414',
    '230304197907184415',
    '230304197304274411',
    '230304197008134430',
    '230306199309094040',
    '230304197308234417',
    '230304197301175071',
    '23030419710422441X',
    '23030519850417431X',
    '230304197510224458',
    '230305198104195023',
    '230302199202105023',
    '230304196906254472',
    '230304197211104413',
    '230303196911296419',
    '230304196810174419',
    '230304197702175419',
    '23030419740427542X',
    '230304197308284828',
    '230304197405024411',
    '230304198309194810',
    '230304197409104224',
    '230304197507214814',
    '23030419800211402X',
    '230307198108014018',
    '230304198302134428',
    '230304199509284612',
    '230303199008056412',
    '230304198702014628',
    '230304196901144813',
    '230304197712114820',
    '230304198704204425',
    '230304198308124810',
    '230304199205044433',
    '23030419741114441X',
    '230304197003314416',
    '230304198507144822',
    '232103198310080938',
    '23030419800125424X',
    '23030419711001462X',
    '230304198808014829',
    '230304197608064819',
    '230304198304244225',
    '230304197104055417',
    '230621197411141561',
    '230303199901076627',
    '23030419810707401X',
    '230304199005074419',
    '230304198002234224',
    '23030419760817441X',
    '230304197305185023',
    '230304199003224428',
    '231026197710152113',
    '23030419790822444X',
    '230307200008294024',
    '230304199802134445',
    '230304198909264616',
    '230304198706084826',
    '230304197403124814',
    '23030419770509442X',
    '230304197203154410',
    '230304196710134612',
    '230304198609094432']
    lis2 = ['230304197204164426',
    '230304199802204212',
    '23030419720224481X',
    '23030419860117442X',
    '230304199707064418',
    '230304200001124816',
    '230304198902124426',
    '230304197310275429',
    '230304200009214015',
    '23030419960223461X',
    '230304198501184428']
    jb.main(['230304199812134413'])
    
    
